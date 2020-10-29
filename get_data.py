#%%
from typing import Collection
import pandas as pd
import json
import requests
from openpyxl import load_workbook
import time

# %%
def get_token(keyword,geo='',timespace="2020-01-01 2020-10-22"):
  headers = {
    'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36',
    'x-client-data': 'CJe2yQEIprbJAQjBtskBCKmdygEIl6zKAQisx8oBCPbHygEI58jKAQjpyMoBCLTLygEI29XKAQjB18oBGIrBygE=',
    'referer': 'https://trends.google.com/trends/explore?q=pytorch',
    'cookie': '__utmc=10102256; __utma=10102256.31392724.1583402727.1586332529.1586398363.11; __utmz=10102256.1586398363.11.11.utmcsr=shimo.im|utmccn=(referral)|utmcmd=referral|utmcct=/docs/qxW86VTXr8DK6HJX; __utmt=1; __utmb=10102256.9.9.1586398779015; ANID=AHWqTUlRutPWkqC3UpC_-5XoYk6zqoDW3RQX5ePFhLykky73kQ0BpL32ATvqV3O0; CONSENT=WP.284bc1; NID=202=xLozp9-VAAGa2d3d9-cqyqmRjW9nu1zmK0j50IM4pdzJ6wpWTO_Z49JN8W0s1OJ8bySeirh7pSMew1WdqRF890iJLX4HQwwvVkRZ7zwsBDxzeHIx8MOWf27jF0mVCxktZX6OmMmSA0txa0zyJ_AJ3i9gmtEdLeopK5BO3X0LWRA; 1P_JAR=2020-4-9-2'
  }

  url = 'https://trends.google.com/trends/api/explore?hl=zh-CN&tz=-480&req={{"comparisonItem":[{{"keyword":"{}","geo":"{}","time":"{}"}}],"category":0,"property":""}}&tz=-480'.format(keyword,geo,timespace)
  # print(url)
  r = requests.get(url, headers=headers,timeout=15)
  # print(r.text[:])
  # print(r.text[5:])
  data = json.loads(r.text[5:])
  # print(data)
  req = data['widgets'][0]['request']
  # print(req)
  token = data['widgets'][0]['token']
  # print(token)
  result = {'req':req,'token':token}
  return result


# %%
def TimeTrans(timenum):
    timeStamp = timenum
    timeArray = time.localtime(timeStamp)
    otherStyletime = time.strftime("%Y-%M-%D",timeArray)
    return otherStyletime

def google(keyword,timespace="2020-05-01 2020-10-22"):
    info = get_token(keyword,timespace=timespace)
    req = info['req']
    token = info['token']
    url = 'https://trends.google.com/trends/api/widgetdata/multiline?hl=zh-CN&tz=-480&req={}&token={}&tz=-480'.format(req, token)
    r = requests.get(url)
    print(r.text)

    if r.status_code == 200:
        data = json.loads(r.text.encode().decode('unicode_escape')[6:])['default']['timelineData']
        print(data)
        timestamp = []
        value = []
        keyword = keyword

        for data_e in data:
            timestampstr = int(data_e['time'])
            timestampstr = TimeTrans(timestampstr)
            valuenum = data_e['value'][0]
            print(timestamp,value,keyword)
            timestamp.append(timestampstr)
            value.append(valuenum)

        data_trend = {'time':timestamp,keyword:value}
        print(data_trend)
        # writeDataToExcleFile(data_trend,j)
        return data_trend
def writeDataToExcleFile(data_trend,j):
    wk = load_workbook('datadata.xlsx')
    wk_name = wk.sheetnames
    wk_sheet = wk[wk_name[0]]

    key1name = list(data_trend.keys())[0]
    key2name = list(data_trend.keys())[1]
    print(key1name,key2name)

    for i in range(len(data_trend[key1name])):
        if j==1:
            wk_sheet.cell(row=1,column=1,value=key1name)
            wk_sheet.cell(row=i+1,column=1,value=data_trend[key1name][i])
        wk_sheet.cell(row=1,column=j+1,value=key2name)
        wk_sheet.cell(row=i+1,column=j+1,value=data_trend[key2name][i])
    
    wk.save('datadata.xlsx')

# %%
def main(kw_list,timespace):
    time = google(kw_list[0],timespace)['time']
    dk = {'time':time}
    for kw in kw_list:
        dk[kw] = google(kw,timespace)[kw]
    df = pd.DataFrame(dk).set_index('time')        
    return df

if __name__ == '__main__':
    # kw_list = ['易烊千玺','王俊凯','王源']
    # kw_list = ['Roboco','yuzuki choco','ookami mio','nekomata okayu','shirakami fubuki','minato aqua','natsuiro matsuri','akai haato','kiryu coco','Tokoyami Towa','Hoshimachi Suisei','Sakura Miko']

    timespace = '2020-05-01 2020-10-22'
    df = main(kw_list,timespace)
    print(df.tail())

#%%
# import bar_chart_race as bcr
import bar_chart_race as bcr
bcr.bar_chart_race(df.cumsum(),filename='test.mp4',title='谷歌趋势',fixed_max=False,period_length=1000,n_bars=6)


